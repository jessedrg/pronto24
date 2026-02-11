import openpyxl

wb = openpyxl.load_workbook('/Users/jesse.dragstra/Documents/pronto24/hogarya.eu-Performance-on-Search-2026-02-11.xlsx')

ws = wb['Consultas']
queries = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    q, clicks, impr, ctr, pos = row
    queries.append({'q': str(q or ''), 'clicks': int(clicks or 0), 'impr': int(impr or 0), 'pos': float(pos or 0)})

prof_map = {'electricista': 'electricista', 'fontaner': 'fontanero', 'cerrajer': 'cerrajero', 'desatasc': 'desatascos', 'caldera': 'calderas'}
prof_counts = {}
for q in queries:
    ql = q['q'].lower()
    for key, name in prof_map.items():
        if key in ql:
            prof_counts.setdefault(name, {'clicks': 0, 'impr': 0, 'count': 0})
            prof_counts[name]['clicks'] += q['clicks']
            prof_counts[name]['impr'] += q['impr']
            prof_counts[name]['count'] += 1
            break

print('=== BREAKDOWN BY PROFESSION ===')
for p, v in sorted(prof_counts.items(), key=lambda x: x[1]['clicks'], reverse=True):
    print(f"{p}: {v['count']} queries, {v['clicks']} clicks, {v['impr']} impressions")

total_clicks = sum(q['clicks'] for q in queries)
total_impr = sum(q['impr'] for q in queries)
print(f"\nTOTAL: {total_clicks} clicks, {total_impr} impressions across {len(queries)} queries")

print("\n=== TOP 40 QUERIES BY CLICKS ===")
by_clicks = sorted(queries, key=lambda x: x['clicks'], reverse=True)
for q in by_clicks[:40]:
    print(f"{q['clicks']}c {q['impr']}i pos{q['pos']:.1f} | {q['q']}")

print("\n=== TOP 30 QUERIES BY IMPRESSIONS ===")
by_impr = sorted(queries, key=lambda x: x['impr'], reverse=True)
for q in by_impr[:30]:
    print(f"{q['clicks']}c {q['impr']}i pos{q['pos']:.1f} | {q['q']}")

# Pages
ws2 = wb['Paginas'] if 'Paginas' in wb.sheetnames else wb['Páginas']
pages = []
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    if i == 0:
        continue
    url, clicks, impr, ctr, pos = row
    pages.append({'url': str(url or ''), 'clicks': int(clicks or 0), 'impr': int(impr or 0), 'pos': float(pos or 0)})

print("\n=== TOP 30 PAGES BY CLICKS ===")
by_clicks_p = sorted(pages, key=lambda x: x['clicks'], reverse=True)
for p in by_clicks_p[:30]:
    print(f"{p['clicks']}c {p['impr']}i pos{p['pos']:.1f} | {p['url']}")

# Devices
ws3 = wb['Dispositivos']
print("\n=== DEVICES ===")
for i, row in enumerate(ws3.iter_rows(values_only=True)):
    if i == 0:
        continue
    print(row)

# Daily trend
ws4 = wb['Gráfico'] if 'Gráfico' in wb.sheetnames else wb['Grafico']
print("\n=== DAILY TREND (last 10 days) ===")
rows_chart = list(ws4.iter_rows(values_only=True))
for row in rows_chart[-10:]:
    print(row)
